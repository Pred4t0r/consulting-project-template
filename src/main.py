"""Main module for MLS Executive Studio - Real Estate Property Analysis Tool.

This module provides functionality for extracting real estate listing data from various
sources, calculating investment metrics, and generating executive reports in Excel format.
"""

from __future__ import annotations

import io
import json
import re
import unicodedata
from copy import deepcopy
from dataclasses import asdict, dataclass
from datetime import datetime
from typing import Any, Callable
from urllib.parse import parse_qs, quote_plus, urlparse
from xml.etree import ElementTree as ET

import pandas as pd
import requests
import streamlit as st
from bs4 import BeautifulSoup, Tag
from openpyxl.cell.cell import MergedCell
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill


# Web request configuration
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    )
}
TIMEOUT_S = 18  # Timeout for general requests
SEARCH_TIMEOUT_S = 6  # Timeout for search requests

# Search configuration
MAX_CANDIDATES = 20  # Maximum number of candidate URLs to collect
MAX_SEARCH_QUERIES = 5  # Maximum number of search queries to execute
SEARCH_PROVIDERS = ("bing_rss", "duckduckgo_html", "duckduckgo_lite")

# Domain hints for scoring URLs
REAL_ESTATE_DOMAIN_HINTS = (
    "realtor.com",
    "zillow.com",
    "redfin.com",
    "homes.com",
    "compass.com",
    "coldwellbankerhomes.com",
    "remax.com",
    "century21.com",
    "trulia.com",
    "movoto.com",
    "bhhs.com",
    "estately.com",
    "propertyshark.com",
    "landwatch.com",
    "smartmls",
)

NOISE_DOMAIN_HINTS = (
    "google.com",
    "youtube.com",
    "facebook.com",
    "instagram.com",
    "whatsapp.com",
    "ae.com",
    "amazon.com",
    "support.google.com",
    "microsoft.com",
    "zhihu.com",
    "chip.de",
)


@dataclass
class PropertyRecord:
    """Data class representing a real estate property record with various attributes."""
    
    # Required fields
    mls_number: str  # MLS listing number
    state: str  # State where the property is located
    
    # Optional fields
    listing_url: str | None = None  # URL of the property listing
    address: str | None = None  # Property address
    city: str | None = None  # City where the property is located
    zip_code: str | None = None  # ZIP code of the property
    price: float | None = None  # Listing price of the property
    bedrooms: float | None = None  # Number of bedrooms
    bathrooms: float | None = None  # Number of bathrooms
    living_area_sqft: float | None = None  # Living area in square feet
    lot_size_sqft: float | None = None  # Lot size in square feet
    property_type: str | None = None  # Type of property (e.g., house, condo)
    year_built: int | None = None  # Year the property was built
    broker_name: str | None = None  # Name of the listing broker
    source_name: str | None = None  # Name of the source where the listing was found
    photo_url: str | None = None  # URL of the property photo


@dataclass
class ExecutiveMetrics:
    """Data class representing calculated executive metrics for property analysis."""
    
    price_per_sqft: float | None  # Price per square foot
    rent_proxy_monthly: float | None  # Estimated monthly rent
    estimated_noi: float | None  # Estimated Net Operating Income
    estimated_cap_rate: float | None  # Estimated capitalization rate
    recommendation: str  # Recommendation based on metrics


@dataclass
class SearchAttempt:
    """Data class representing a search attempt for property listings."""
    
    provider: str  # Search provider used
    query: str  # Query string used for the search
    status_code: int | None  # HTTP status code of the search request
    outcome: str  # Outcome of the search attempt
    hits: int = 0  # Number of hits returned by the search
    detail: str | None = None  # Additional details about the search attempt


def _to_float(value: Any) -> float | None:
    """Convert a value to float, handling various input types.
    
    Args:
        value: The value to convert to float
        
    Returns:
        The float representation of the value, or None if conversion fails
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value)
    match = re.search(r"-?[\d,.]+", text)
    if not match:
        return None
    try:
        return float(match.group(0).replace(",", ""))
    except ValueError:
        return None


def _decode_search_result_href(href: str | None) -> str | None:
    """Decode search result href, handling redirects and protocol fixes.
    
    Args:
        href: The href to decode
        
    Returns:
        The decoded href, or None if it's invalid
    """
    if not href:
        return None
    if href.startswith("//"):
        href = f"https:{href}"

    # DuckDuckGo HTML results often wrap targets in a redirect URL.
    if "duckduckgo.com/l/" in href:
        parsed = urlparse(href)
        target = parse_qs(parsed.query).get("uddg", [None])[0]
        return target or None
    return href if href.startswith("http") else None


def _is_blocked_search_response(status_code: int, html: str) -> bool:
    """Check if a search response indicates the request was blocked.
    
    Args:
        status_code: HTTP status code of the response
        html: HTML content of the response
        
    Returns:
        True if the response indicates the request was blocked, False otherwise
    """
    if status_code in {202, 403, 429}:
        return True
    text = html.lower()
    markers = (
        "verify you are human",
        "unusual traffic",
        "automated traffic",
        "/sorry/index",
    )
    return any(marker in text for marker in markers)


def _build_search_queries(mls_number: str, state: str) -> list[str]:
    """Build search queries for finding property listings.
    
    Args:
        mls_number: The MLS number to search for
        state: The state to search in
        
    Returns:
        A list of search query strings
    """
    state_fragment = state if state and state != "Other" else ""
    base_terms = [
        f'"{mls_number}" {state_fragment} MLS listing',
        f'{mls_number} {state_fragment} MLS listing',
        f'"MLS {mls_number}" {state_fragment} real estate',
        f'"{mls_number}" "smartmls"',
        f'"{mls_number}" site:compass.com',
        f'"{mls_number}" site:coldwellbankerhomes.com',
        f'"{mls_number}" site:realtor.com',
        f'"{mls_number}" site:homes.com',
        f'"{mls_number}" site:zillow.com',
        f'"{mls_number}" site:redfin.com',
    ]
    # Keep order stable and remove blanks/duplicates.
    out: list[str] = []
    for q in base_terms:
        normalized = " ".join(q.split()).strip()
        if normalized and normalized not in out:
            out.append(normalized)
    return out


def _has_mls_match(page_html: str, page_text: str, url: str, mls_number: str) -> bool:
    """Check if the page contains a match for the MLS number.
    
    Args:
        page_html: HTML content of the page
        page_text: Text content of the page
        url: URL of the page
        mls_number: MLS number to search for
        
    Returns:
        True if the page contains a match for the MLS number, False otherwise
    """
    mls_escaped = re.escape(mls_number.strip())
    patterns = [
        rf"\bMLS\s*(?:#|ID|Number|No\.?)?\s*[:#]?\s*{mls_escaped}\b",
        rf"\b{mls_escaped}\b",
    ]
    for pattern in patterns:
        if re.search(pattern, page_text, re.IGNORECASE):
            return True
        if re.search(pattern, page_html, re.IGNORECASE):
            return True
    return mls_number in url


def _candidate_url_score(url: str) -> int:
    """Score a URL based on its likelihood of containing real estate listings.
    
    Args:
        url: The URL to score
        
    Returns:
        A score representing the likelihood of the URL containing real estate listings
    """
    parsed = urlparse(url)
    host = parsed.netloc.lower()
    path = parsed.path.lower()
    score = 0
    if any(hint in host for hint in REAL_ESTATE_DOMAIN_HINTS):
        score += 10
    if any(hint in host for hint in NOISE_DOMAIN_HINTS):
        score -= 10
    if any(term in path for term in ("listing", "property", "home", "real-estate", "realestate")):
        score += 2
    if any(token in url.lower() for token in ("mls", "smartmls")):
        score += 3
    return score


def _parse_duckduckgo_html_links(html: str) -> list[str]:
    """Parse links from DuckDuckGo HTML search results.
    
    Args:
        html: HTML content of the search results page
        
    Returns:
        A list of URLs extracted from the search results
    """
    soup = BeautifulSoup(html, "html.parser")
    links: list[str] = []
    for a in soup.select("a.result__a"):
        href = _decode_search_result_href(a.get("href"))
        if href and href not in links:
            links.append(href)
    return links


def _parse_duckduckgo_lite_links(html: str) -> list[str]:
    """Parse links from DuckDuckGo Lite search results.
    
    Args:
        html: HTML content of the search results page
        
    Returns:
        A list of URLs extracted from the search results
    """
    soup = BeautifulSoup(html, "html.parser")
    links: list[str] = []
    for a in soup.select("a"):
        href = _decode_search_result_href(a.get("href"))
        if not href:
            continue
        if "duckduckgo.com" in urlparse(href).netloc.lower():
            continue
        if href not in links:
            links.append(href)
    return links


def _parse_bing_html_links(html: str) -> list[str]:
    """Parse links from Bing HTML search results.
    
    Args:
        html: HTML content of the search results page
        
    Returns:
        A list of URLs extracted from the search results
    """
    soup = BeautifulSoup(html, "html.parser")
    links: list[str] = []
    for selector in ("li.b_algo h2 a", "main a"):
        for a in soup.select(selector):
            href = a.get("href")
            if not href or not href.startswith("http"):
                continue
            host = urlparse(href).netloc.lower()
            if "bing.com" in host or "microsoft.com" in host:
                continue
            if href not in links:
                links.append(href)
        if links:
            break
    return links


def _parse_bing_rss_links(xml_text: str) -> list[str]:
    """Parse links from Bing RSS search results.
    
    Args:
        xml_text: XML content of the RSS feed
        
    Returns:
        A list of URLs extracted from the RSS feed
    """
    links: list[str] = []
    try:
        root = ET.fromstring(xml_text)
    except ET.ParseError:
        return links

    for item in root.findall("./channel/item"):
        link_node = item.find("link")
        if link_node is None or not link_node.text:
            continue
        href = link_node.text.strip()
        if href.startswith("http") and href not in links:
            links.append(href)
    return links


def _fetch_search_results(query_text: str, provider: str) -> tuple[list[str], SearchAttempt]:
    """Fetch search results from a specific provider.
    
    Args:
        query_text: The search query text
        provider: The search provider to use
        
    Returns:
        A tuple containing a list of URLs and a SearchAttempt object with metadata
    """
    query = quote_plus(query_text)
    if provider == "duckduckgo_html":
        url = f"https://duckduckgo.com/html/?q={query}"
        parser = _parse_duckduckgo_html_links
    elif provider == "duckduckgo_lite":
        url = f"https://lite.duckduckgo.com/lite/?q={query}"
        parser = _parse_duckduckgo_lite_links
    elif provider == "bing_rss":
        url = f"https://www.bing.com/search?format=rss&q={query}"
        parser = _parse_bing_rss_links
    elif provider == "bing_html":
        url = f"https://www.bing.com/search?q={query}"
        parser = _parse_bing_html_links
    else:
        return [], SearchAttempt(provider, query_text, None, "error", detail="unknown provider")

    try:
        response = requests.get(url, headers=HEADERS, timeout=SEARCH_TIMEOUT_S)
    except requests.RequestException as exc:
        return [], SearchAttempt(provider, query_text, None, "network_error", detail=str(exc))

    if _is_blocked_search_response(response.status_code, response.text):
        return [], SearchAttempt(provider, query_text, response.status_code, "blocked")
    if response.status_code != 200:
        return [], SearchAttempt(provider, query_text, response.status_code, "http_error")

    links = parser(response.text)
    return links, SearchAttempt(
        provider,
        query_text,
        response.status_code,
        "ok" if links else "no_results",
        hits=len(links),
    )


def discover_listing_candidates(
    mls_number: str,
    state: str,
    reporter: Callable[[str], None] | None = None,
) -> tuple[list[str], list[SearchAttempt]]:
    """Discover potential listing URLs for a given MLS number and state.
    
    Args:
        mls_number: The MLS number to search for
        state: The state to search in
        reporter: Optional callback function to report progress
        
    Returns:
        A tuple containing a list of candidate URLs and a list of search attempts
    """
    links: list[str] = []
    attempts: list[SearchAttempt] = []
    blocked_providers: set[str] = set()
    for query_text in _build_search_queries(mls_number, state)[:MAX_SEARCH_QUERIES]:
        for provider in SEARCH_PROVIDERS:
            if provider in blocked_providers:
                continue
            provider_links, attempt = _fetch_search_results(query_text, provider)
            attempts.append(attempt)
            if reporter is not None:
                reporter(
                    f"search {provider} | query={query_text} | outcome={attempt.outcome} "
                    f"| status={attempt.status_code} | hits={attempt.hits}"
                )
            if attempt.outcome == "blocked":
                blocked_providers.add(provider)
            ranked_links = sorted(provider_links, key=_candidate_url_score, reverse=True)
            for href in ranked_links:
                score = _candidate_url_score(href)
                if provider == "bing_rss" and score <= 0:
                    continue
                if score < 0:
                    continue
                if href not in links:
                    links.append(href)
                    if reporter is not None:
                        reporter(f"candidate + score={score} | {href}")
                if len(links) >= MAX_CANDIDATES:
                    return links, attempts
    return links, attempts


def diagnose_url_access(url: str) -> str | None:
    """Diagnose access issues for a given URL.
    
    Args:
        url: The URL to diagnose
        
    Returns:
        A string describing the access issue, or None if no issues found
    """
    try:
        response = requests.get(url, headers=HEADERS, timeout=SEARCH_TIMEOUT_S)
    except requests.RequestException as exc:
        return f"Could not fetch the manual URL ({type(exc).__name__})."

    text_lower = response.text.lower()
    if response.status_code in {401, 403, 429}:
        if "just a moment" in text_lower or "cf-challenge" in text_lower or "cloudflare" in text_lower:
            return (
                f"The listing site blocked automated access (HTTP {response.status_code}, anti-bot challenge). "
                "The URL matches, but this app cannot read the page content from a normal requests call."
            )
        return f"The listing site blocked automated access (HTTP {response.status_code})."
    if response.status_code >= 400:
        return f"The listing page returned HTTP {response.status_code}."
    return None


def search_listing_candidates(mls_number: str, state: str) -> list[str]:
    """Search for listing candidates using the provided MLS number and state.
    
    Args:
        mls_number: The MLS number to search for
        state: The state to search in
        
    Returns:
        A list of candidate URLs
    """
    links, _attempts = discover_listing_candidates(mls_number, state)
    return links


def _extract_json_ld(soup: BeautifulSoup) -> list[dict[str, Any]]:
    """Extract JSON-LD structured data from the HTML soup.
    
    Args:
        soup: BeautifulSoup object representing the HTML content
        
    Returns:
        A list of dictionaries containing the extracted JSON-LD data
    """
    docs: list[dict[str, Any]] = []
    for node in soup.find_all("script", {"type": "application/ld+json"}):
        if not isinstance(node, Tag):
            continue
        raw = ((node.string if node.string is not None else node.get_text()) or "").strip()
        if not raw:
            continue
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, list):
                docs.extend([x for x in parsed if isinstance(x, dict)])
            elif isinstance(parsed, dict):
                docs.append(parsed)
        except json.JSONDecodeError:
            continue
    return docs


def _first_image_url(value: Any) -> str | None:
    """Extract the first valid image URL from a value that could be a string, list, or dict.
    
    Args:
        value: The value to extract an image URL from
        
    Returns:
        The first valid image URL found, or None if none found
    """
    if isinstance(value, str) and value.startswith("http"):
        return value
    if isinstance(value, list):
        for item in value:
            found = _first_image_url(item)
            if found:
                return found
    if isinstance(value, dict):
        for key in ("url", "contentUrl"):
            v = value.get(key)
            if isinstance(v, str) and v.startswith("http"):
                return v
    return None


def _extract_from_html_content(
    url: str,
    html_content: str,
    mls_number: str,
    state: str,
) -> PropertyRecord | None:
    """Extract property record data from HTML content.
    
    Args:
        url: The URL of the page being parsed
        html_content: The HTML content to parse
        mls_number: The MLS number to look for
        state: The state where the property is located
        
    Returns:
        A PropertyRecord object with the extracted data, or None if extraction failed
    """
    soup = BeautifulSoup(html_content, "html.parser")
    full_text = soup.get_text(" ", strip=True)
    mls_match = _has_mls_match(html_content, full_text, url, mls_number)

    record = PropertyRecord(mls_number=mls_number, state=state, listing_url=url)
    docs = _extract_json_ld(soup)

    for doc in docs:
        dtype = str(doc.get("@type", "")).lower()
        if "residence" in dtype or "house" in dtype or "singlefamilyresidence" in dtype:
            record.address = (
                doc.get("name")
                or (doc.get("address") or {}).get("streetAddress")
                or record.address
            )
            adr = doc.get("address") or {}
            record.city = adr.get("addressLocality") or record.city
            record.zip_code = adr.get("postalCode") or record.zip_code
            record.bedrooms = _to_float(doc.get("numberOfRooms") or doc.get("numberOfBedrooms") or record.bedrooms)
            record.bathrooms = _to_float(doc.get("numberOfBathroomsTotal") or doc.get("numberOfBathrooms") or record.bathrooms)
            record.living_area_sqft = _to_float((doc.get("floorSize") or {}).get("value") or record.living_area_sqft)
            record.property_type = doc.get("@type") or record.property_type
            record.year_built = int(_to_float(doc.get("yearBuilt")) or 0) or record.year_built
            record.photo_url = _first_image_url(doc.get("image")) or record.photo_url

        offers = doc.get("offers")
        if isinstance(offers, dict):
            record.price = _to_float(offers.get("price") or record.price)

        if "realestateagent" in dtype:
            record.broker_name = doc.get("name") or record.broker_name

        if record.photo_url is None:
            record.photo_url = _first_image_url(doc.get("image")) or record.photo_url

    if not record.price:
        m = re.search(r"\$\s?([\d,]+)", full_text)
        record.price = _to_float(m.group(1)) if m else None
    if not record.bedrooms:
        m = re.search(r"(\d+(?:\.\d+)?)\s*(?:bed|beds|bedroom)", full_text, re.IGNORECASE)
        record.bedrooms = _to_float(m.group(1)) if m else None
    if not record.bathrooms:
        m = re.search(r"(\d+(?:\.\d+)?)\s*(?:bath|baths|bathroom)", full_text, re.IGNORECASE)
        record.bathrooms = _to_float(m.group(1)) if m else None
    if not record.living_area_sqft:
        m = re.search(r"([\d,]+)\s*(?:sq\.?\s?ft|square feet)", full_text, re.IGNORECASE)
        record.living_area_sqft = _to_float(m.group(1)) if m else None

    title = soup.title.string.strip() if soup.title and soup.title.string else ""
    record.source_name = title[:80] if title else None
    if not record.photo_url:
        meta_img = soup.find("meta", attrs={"property": "og:image"}) or soup.find("meta", attrs={"name": "og:image"})
        if isinstance(meta_img, Tag):
            content = meta_img.get("content")
            if isinstance(content, str) and content.startswith("http"):
                record.photo_url = content

    # Keep a safety check, but avoid false negatives from JS-heavy pages that hide MLS in rendered text.
    if not mls_match and not any([record.address, record.price, record.bedrooms, record.bathrooms]):
        return None
    return record


def extract_from_listing(url: str, mls_number: str, state: str) -> PropertyRecord | None:
    """Extract property record from a listing URL.
    
    Args:
        url: The URL of the listing to extract from
        mls_number: The MLS number to look for
        state: The state where the property is located
        
    Returns:
        A PropertyRecord object with the extracted data, or None if extraction failed
    """
    try:
        response = requests.get(url, headers=HEADERS, timeout=TIMEOUT_S)
        response.raise_for_status()
    except Exception:
        return None
    return _extract_from_html_content(url, response.text, mls_number, state)


def extract_from_pasted_content(
    source_url: str,
    pasted_content: str,
    mls_number: str,
    state: str,
) -> PropertyRecord | None:
    """Extract property record from pasted HTML content.
    
    Args:
        source_url: The URL of the source page
        pasted_content: The HTML content that was pasted
        mls_number: The MLS number to look for
        state: The state where the property is located
        
    Returns:
        A PropertyRecord object with the extracted data, or None if extraction failed
    """
    if not pasted_content.strip():
        return None
    return _extract_from_html_content(source_url.strip() or "manual://pasted-content", pasted_content, mls_number, state)


def calculate_metrics(record: PropertyRecord) -> ExecutiveMetrics:
    """Calculate executive metrics for a property record.
    
    Args:
        record: The PropertyRecord to calculate metrics for
        
    Returns:
        An ExecutiveMetrics object with the calculated metrics
    """
    ppsf = (record.price / record.living_area_sqft) if record.price and record.living_area_sqft else None
    rent = (record.price * 0.0065 / 12) if record.price else None
    noi = (rent * 12 * 0.92 * 0.68) if rent else None
    cap = (noi / record.price) if noi and record.price else None

    score = 0
    if cap and cap >= 0.05:
        score += 1
    if ppsf and ppsf <= 350:
        score += 1
    if record.bedrooms and record.bedrooms >= 3:
        score += 1
    recommendation = "STRONG FIT" if score >= 2 else "REVIEW MANUALLY"
    return ExecutiveMetrics(ppsf, rent, noi, cap, recommendation)


def render_default_workbook(record: PropertyRecord, metrics: ExecutiveMetrics) -> bytes:
    """Render a default workbook with the property record and metrics.
    
    Args:
        record: The PropertyRecord to include in the workbook
        metrics: The ExecutiveMetrics to include in the workbook
        
    Returns:
        Bytes representing the Excel workbook
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Executive File"
    ws["A1"] = "Professional Executive Property File"
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1E3A5F")

    rows = [
        ("Generated At", datetime.utcnow().isoformat(timespec="seconds") + "Z"),
        ("MLS Number", record.mls_number),
        ("State", record.state),
        ("Address", record.address),
        ("City", record.city),
        ("ZIP", record.zip_code),
        ("List Price", record.price),
        ("Beds", record.bedrooms),
        ("Baths", record.bathrooms),
        ("Living Area (sqft)", record.living_area_sqft),
        ("Lot Size (sqft)", record.lot_size_sqft),
        ("Property Type", record.property_type),
        ("Year Built", record.year_built),
        ("Broker", record.broker_name),
        ("Source", record.listing_url),
        ("Price / sqft", metrics.price_per_sqft),
        ("Estimated NOI", metrics.estimated_noi),
        ("Estimated Cap Rate", metrics.estimated_cap_rate),
        ("Recommendation", metrics.recommendation),
    ]
    for idx, (k, v) in enumerate(rows, start=3):
        ws[f"A{idx}"] = k
        ws[f"A{idx}"].font = Font(bold=True)
        ws[f"B{idx}"] = v
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 60

    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def _executive_rows(record: PropertyRecord, metrics: ExecutiveMetrics) -> list[tuple[str, Any]]:
    """Get the rows for the executive summary.
    
    Args:
        record: The PropertyRecord to include in the summary
        metrics: The ExecutiveMetrics to include in the summary
        
    Returns:
        A list of tuples representing the rows in the executive summary
    """
    return [
        ("Generated At", datetime.utcnow().isoformat(timespec="seconds") + "Z"),
        ("MLS Number", record.mls_number),
        ("State", record.state),
        ("Address", record.address),
        ("City", record.city),
        ("ZIP", record.zip_code),
        ("List Price", record.price),
        ("Beds", record.bedrooms),
        ("Baths", record.bathrooms),
        ("Living Area (sqft)", record.living_area_sqft),
        ("Lot Size (sqft)", record.lot_size_sqft),
        ("Property Type", record.property_type),
        ("Year Built", record.year_built),
        ("Broker", record.broker_name),
        ("Source", record.listing_url),
        ("Price / sqft", metrics.price_per_sqft),
        ("Estimated NOI", metrics.estimated_noi),
        ("Estimated Cap Rate", metrics.estimated_cap_rate),
        ("Recommendation", metrics.recommendation),
    ]


def _write_executive_rows_to_sheet(ws: Any, rows: list[tuple[str, Any]], title: str = "Extracted MLS Data") -> None:
    """Write executive rows to a worksheet.
    
    Args:
        ws: The worksheet to write to
        rows: The rows to write
        title: The title for the sheet
    """
    ws["A1"] = title
    ws["A1"].font = Font(size=14, bold=True)
    for idx, (k, v) in enumerate(rows, start=3):
        ws[f"A{idx}"] = k
        ws[f"A{idx}"].font = Font(bold=True)
        ws[f"B{idx}"] = v
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 60


def _is_writable_cell(ws: Any, row: int, col: int) -> bool:
    """Check if a cell in a worksheet is writable (not merged).
    
    Args:
        ws: The worksheet to check
        row: The row number of the cell
        col: The column number of the cell
        
    Returns:
        True if the cell is writable, False otherwise
    """
    return not isinstance(ws.cell(row=row, column=col), MergedCell)


def _safe_write_nearby(ws: Any, row: int, start_col: int, value: Any, max_offset: int = 6) -> bool:
    """Safely write a value to a nearby writable cell.
    
    Args:
        ws: The worksheet to write to
        row: The row number to start searching from
        start_col: The column number to start searching from
        value: The value to write
        max_offset: The maximum number of columns to search ahead
        
    Returns:
        True if the value was written successfully, False otherwise
    """
    for offset in range(max_offset + 1):
        col = start_col + offset
        if _is_writable_cell(ws, row, col):
            ws.cell(row=row, column=col, value=value)
            return True
    return False


def _norm_key(text: str) -> str:
    """Normalize a text string for use as a key.
    
    Args:
        text: The text to normalize
        
    Returns:
        The normalized text string
    """
    ascii_text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]", "", ascii_text.lower())


def _download_photo_bytes(photo_url: str | None) -> bytes | None:
    """Download photo bytes from a URL.
    
    Args:
        photo_url: The URL of the photo to download
        
    Returns:
        The photo bytes if download was successful, None otherwise
    """
    if not photo_url:
        return None
    try:
        r = requests.get(photo_url, headers=HEADERS, timeout=TIMEOUT_S)
        r.raise_for_status()
    except requests.RequestException:
        return None
    content_type = (r.headers.get("Content-Type") or "").lower()
    if not (content_type.startswith("image/") or photo_url.lower().endswith((".jpg", ".jpeg", ".png", ".webp"))):
        return None
    return r.content


def _replace_template_images(wb: Any, photo_bytes: bytes | None) -> None:
    """Replace images in a workbook with a new photo.
    
    Args:
        wb: The workbook to modify
        photo_bytes: The new photo bytes to use
    """
    for ws in wb.worksheets[:3]:
        existing_images = list(getattr(ws, "_images", []))
        if not existing_images:
            continue

        anchors_and_sizes: list[tuple[Any, float, float]] = []
        for img in existing_images:
            anchors_and_sizes.append((deepcopy(getattr(img, "anchor", None)), float(getattr(img, "width", 0)), float(getattr(img, "height", 0))))

        # Remove old/sample images from the template.
        ws._images = []

        if not photo_bytes:
            continue

        # Add the fetched property photo at the same anchors/sizes.
        for anchor, width, height in anchors_and_sizes:
            try:
                new_img = XLImage(io.BytesIO(photo_bytes))
                if width:
                    new_img.width = width
                if height:
                    new_img.height = height
                if anchor is not None:
                    new_img.anchor = anchor
                ws.add_image(new_img)
            except Exception:
                continue


def apply_to_template(template_bytes: bytes, record: PropertyRecord, metrics: ExecutiveMetrics) -> bytes:
    """Apply property record and metrics to an Excel template.
    
    Args:
        template_bytes: The Excel template as bytes
        record: The PropertyRecord to apply to the template
        metrics: The ExecutiveMetrics to apply to the template
        
    Returns:
        The modified Excel file as bytes
    """
    wb = load_workbook(io.BytesIO(template_bytes))
    _replace_template_images(wb, _download_photo_bytes(record.photo_url))
    payload = {**asdict(record), **asdict(metrics), "generated_at": datetime.utcnow().isoformat(timespec="seconds") + "Z"}
    alias_payload: dict[str, Any] = {}
    alias_groups = {
        "generated_at": ["fechageneracion", "generadoel", "fechadegeneracion"],
        "mls_number": ["mls", "mlsnumber", "mlsid", "listingid", "numero mls", "numero de mls"],
        "state": ["estado", "provincia"],
        "listing_url": ["source", "sourceurl", "listingurl", "url", "fuente", "enlace", "link"],
        "address": ["direccion", "domicilio", "addressline"],
        "city": ["ciudad", "municipio"],
        "zip_code": ["zip", "zipcode", "postalcode", "codigo postal", "cp"],
        "price": ["listprice", "saleprice", "askingprice", "precio", "precio lista", "precio de venta"],
        "bedrooms": ["beds", "bed", "bedroomcount", "habitaciones", "recamaras", "dormitorios"],
        "bathrooms": ["baths", "bath", "bathroomcount", "banos", "baños", "servicios"],
        "living_area_sqft": ["sqft", "livingarea", "livingareasqft", "squarefeet", "superficie", "area", "metros cuadrados", "m2"],
        "lot_size_sqft": ["lotsize", "tamano lote", "tamano del lote", "superficie lote"],
        "property_type": ["propertytype", "tipo propiedad", "tipo de propiedad"],
        "year_built": ["yearbuilt", "ano construido", "año construido", "anio construccion", "año de construccion"],
        "broker_name": ["broker", "agente", "inmobiliaria", "corredor"],
        "source_name": ["nombre fuente", "sitio fuente"],
        "price_per_sqft": ["ppsf", "pricepersqft"],
        "rent_proxy_monthly": ["rent", "renta", "renta mensual", "alquiler mensual"],
        "estimated_cap_rate": ["caprate", "estimatedcaprate", "tasa capitalizacion", "cap rate"],
        "estimated_noi": ["noi", "estimatednoi", "ingreso operativo neto"],
        "recommendation": ["recommendation", "recomendacion"],
    }
    for key, value in payload.items():
        key_norm = _norm_key(key)
        if key_norm:
            alias_payload[key_norm] = value
    for canonical_key, aliases in alias_groups.items():
        if canonical_key in payload:
            for alias in aliases:
                alias_payload[_norm_key(alias)] = payload[canonical_key]

    # Clear all cells in the target worksheets before populating with new data
    target_sheets = wb.worksheets[:3]
    for ws in target_sheets:
        for row in ws.iter_rows():
            for cell in row:
                # Only replace cells that contain template placeholders or keys
                if not isinstance(cell.value, str):
                    continue
                raw_text = cell.value
                normalized = _norm_key(raw_text)

                # Placeholder replacement in the same cell (e.g. {{mls_number}}, [[price]], <city>).
                compact = _norm_key(raw_text)
                for key_norm, val in alias_payload.items():
                    token_candidates = ("{{" + key_norm + "}}", "[[" + key_norm + "]]", "<" + key_norm + ">")
                    if any(token in raw_text.lower().replace(" ", "") for token in token_candidates):
                        cell.value = "" if val is None else str(val)
                        break
                # If the whole cell is a placeholder-like key, replace the cell itself.
                if normalized in alias_payload:
                    cell.value = alias_payload[normalized]
                    continue

                # Label/value template pattern: write to the next cell when the label references a payload key.
                for key_norm, val in alias_payload.items():
                    if key_norm and key_norm in normalized:
                        if _safe_write_nearby(ws, cell.row, cell.column + 1, val):
                            break

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def build_ui() -> None:
    st.set_page_config(page_title="MLS Executive Studio", page_icon="🏢", layout="wide")
    st.markdown(
        """
        <style>
          .stApp { background: linear-gradient(180deg,#eaf0f7 0%,#f7f9fc 100%); }
          .hero { background:#113355; color:white; padding:20px; border-radius:14px; }
          .sub { color:#d6e3f5; margin-top:8px; }
        </style>
        <div class="hero">
          <h2 style="margin:0;">MLS Executive Studio</h2>
          <p class="sub">SmartMLS-inspired workflow: MLS + State in, decision-ready executive Excel out.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    c1, c2 = st.columns([1, 1])
    with c1:
        mls_number = st.text_input("MLS Number", placeholder="e.g., 24003521")
    with c2:
        state = st.selectbox("State", ["CT", "NY", "NJ", "MA", "FL", "CA", "TX", "Other"])

    manual_url = st.text_input(
        "Listing URL (optional)",
        placeholder="Paste a listing URL to bypass web search if search engines block requests",
    )
    browser_mode = st.checkbox(
        "Browser-assisted mode (paste page source/HTML)",
        value=False,
        help="Use this when listing sites block bots. Open the page in your browser, view page source, and paste it here.",
    )
    pasted_source_url = ""
    pasted_page_content = ""
    if browser_mode:
        pasted_source_url = st.text_input(
            "Browser page URL",
            placeholder="Paste the URL of the page you opened in your browser",
        )
        pasted_page_content = st.text_area(
            "Pasted Page HTML / Source",
            height=220,
            placeholder="Paste page source (Ctrl+U in browser) or copied page HTML here...",
        )
    template_file = st.file_uploader("Upload Excel Template (optional)", type=["xlsx"])

    if st.button("Generate Executive File", type="primary"):
        if not mls_number.strip():
            st.error("MLS number is required.")
            return

        live_debug = st.expander("Live Debug (Real Time)", expanded=True)
        live_status = live_debug.empty()
        live_progress = live_debug.progress(0, text="Waiting to start...")
        live_log = live_debug.empty()
        live_lines: list[str] = []

        def _debug(msg: str) -> None:
            ts = datetime.now().strftime("%H:%M:%S")
            live_lines.append(f"[{ts}] {msg}")
            live_log.code("\n".join(live_lines[-200:]))

        search_attempts: list[SearchAttempt] = []
        with st.spinner("Searching listing sources and extracting verified fields..."):
            live_status.info("Starting search workflow...")
            if browser_mode and pasted_page_content.strip():
                live_progress.progress(10, text="Browser-assisted mode")
                _debug("browser-assisted mode enabled")
                _debug(f"pasted content length={len(pasted_page_content)}")
                source_url = pasted_source_url.strip() or manual_url.strip() or "manual://pasted-content"
                _debug(f"pasted source url | {source_url}")
                candidates = [source_url]
                parsed = []
                live_status.info("Parsing pasted browser content")
                live_progress.progress(70, text="Parsing pasted browser content")
                record = extract_from_pasted_content(source_url, pasted_page_content, mls_number.strip(), state)
                if record is None:
                    _debug("parse no-match/unreadable | pasted browser content")
                else:
                    parsed.append(record)
                    _debug(
                        "parse match | "
                        f"url={source_url} | price={record.price} | beds={record.bedrooms} | baths={record.bathrooms}"
                    )
            elif manual_url.strip():
                candidates = [manual_url.strip()]
                live_progress.progress(10, text="Manual URL mode")
                _debug(f"manual url provided | {manual_url.strip()}")
                parsed = []
            else:
                live_progress.progress(5, text="Searching providers...")
                _debug(f"search start | mls={mls_number.strip()} | state={state}")
                candidates, search_attempts = discover_listing_candidates(mls_number.strip(), state, reporter=_debug)
                parsed = []

            if not parsed:
                if not candidates:
                    _debug("no candidate URLs found")

                total_candidates = max(len(candidates), 1)
                for idx, url in enumerate(candidates, start=1):
                    live_status.info(f"Checking candidate {idx}/{len(candidates)}")
                    parse_pct = 10 + int(85 * idx / total_candidates)
                    live_progress.progress(parse_pct, text=f"Parsing candidate {idx}/{len(candidates)}")
                    _debug(f"parse start | {url}")
                    record = extract_from_listing(url, mls_number.strip(), state)
                    if record is None:
                        _debug(f"parse no-match/unreadable | {url}")
                        continue
                    parsed.append(record)
                    _debug(
                        "parse match | "
                        f"url={url} | price={record.price} | beds={record.bedrooms} | baths={record.bathrooms}"
                    )
                    break
            live_progress.progress(100, text="Done")
            live_status.success("Workflow completed.")

        with st.expander("Search Diagnostics", expanded=not parsed):
            if manual_url.strip():
                st.info("Manual URL provided, so web search was skipped.")
            elif browser_mode and pasted_page_content.strip():
                st.info("Browser-assisted mode provided pasted page content, so web search/fetch was skipped.")
            elif search_attempts:
                st.dataframe(pd.DataFrame([asdict(a) for a in search_attempts]), use_container_width=True)
                blocked_count = sum(1 for a in search_attempts if a.outcome == "blocked")
                if blocked_count:
                    st.warning(
                        f"{blocked_count} search requests were blocked/challenged by search engines. "
                        "Paste a listing URL above for the most reliable fallback."
                    )
            else:
                st.info("No search attempts recorded.")

            if candidates:
                st.write("Candidate URLs")
                st.code("\n".join(candidates[:MAX_CANDIDATES]))

        if not parsed:
            if browser_mode and not pasted_page_content.strip():
                st.error("Browser-assisted mode is enabled, but no page HTML/source was pasted.")
                return
            if browser_mode and pasted_page_content.strip():
                st.error(
                    "Could not parse the pasted browser content for this MLS. "
                    "Use the page source (Ctrl+U) rather than copied visible text, and ensure the MLS number appears in the source."
                )
                return
            if manual_url.strip():
                manual_failure = diagnose_url_access(manual_url.strip())
                if manual_failure:
                    st.error(manual_failure)
                    st.info(
                        "This is not an MLS mismatch. It is a website access restriction. "
                        "Try another listing source URL for the same MLS (Realtor/Zillow/etc.) or a site that does not block bots."
                    )
                    return
            st.error(
                "No listing source with matching MLS number was found. "
                "This can happen when search engines block automated requests. "
                "Try pasting a listing URL in the optional field above."
            )
            return

        record = parsed[0]
        metrics = calculate_metrics(record)

        st.success(f"Found source: {record.listing_url}")
        st.dataframe(pd.DataFrame([asdict(record) | asdict(metrics)]), use_container_width=True)

        if template_file is not None:
            result = apply_to_template(template_file.read(), record, metrics)
            filename = f"executive_file_{mls_number}_{state}_template.xlsx"
        else:
            result = render_default_workbook(record, metrics)
            filename = f"executive_file_{mls_number}_{state}.xlsx"

        st.download_button(
            "Download Executive Excel",
            data=result,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            on_click="ignore",
        )


if __name__ == "__main__":
    build_ui()
